import openpyxl

# Datos de ejemplo
header = ['Total', 'Descuentos', 'Total']
mes = [70000, 5000]
restos = [45000, 15000]

# Crear un nuevo libro de trabajo
workbook = openpyxl.Workbook()
sheet = workbook.active

# Escribir los encabezados en la fila 1
for i in range(len(header)):
    sheet.cell(row=1, column=i+1).value = header[i]

# Escribir los datos de mes y restos en las filas 2 y 3
for i in range(len(mes)):
    sheet.cell(row=i+2, column=1).value = mes[i]
    sheet.cell(row=i+2, column=2).value = restos[i]

# Guardar el archivo
workbook.save("datos_meses.xlsx")