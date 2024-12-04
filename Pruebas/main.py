import openpyxl
import datetime

fecha_actual = datetime.datetime.now()
fecha_f = fecha_actual.strftime("%Y-%m-%d")

try:
    wb = openpyxl.load_workbook(f'Factura{fecha_f}.xlsx')
    ws = wb["Sheet"]
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active

descuentos = [15950, 18333,0]
motivo = ['Hamburguesa Mostaza', '3er cuota de zapatillas','Porcentaje Inflacion:2,7 segun tu amigo milei JAJA']
total = [70000,0,1890]

# CABECERAS
ws.cell(column=1, row=1, value='Descuentos')
ws.cell(column=2, row=1, value='Motivos')
ws.cell(column=3, row=1, value='Totales')


for i in range(2, len(descuentos) + 2):
    ws.cell(column=1, row=i, value=f"{descuentos[i-2]:,.2f}")
    ws.cell(column=2, row=i, value=motivo[i-2])
    ws.cell(column=3, row=i, value=f"{total[i-2]:,.2f}")

# Calcula el total de todos los descuentos

total_FINAL = sum(total) - sum(descuentos)

ws.cell(column=1, row=len(descuentos) + 3, value="Total Teorico")
ws.cell(column=3, row=len(descuentos) + 3, value=total_FINAL)

    
wb.save(f'Factura{fecha_f}.xlsx')
